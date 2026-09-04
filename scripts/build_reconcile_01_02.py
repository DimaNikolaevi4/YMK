# -*- coding: utf-8 -*-
"""Сводная таблица сверки часов МДК 01.02 М-21: нагрузка / УП / РП / КТП.
Источники: нагрузка_2026-2027.md, учебные_планы/исходники/08.02.09_УП_2025.doc,
RP/08.02.09_ПМ.01_РП_2024.doc, KTP/КТП МДК 01.02 М-21.docx."""
import sys, os

XLSX_SKILL_DIR = '/home/z/my-project/skills/xlsx'
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, 'templates')]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from base import (setup_sheet, style_header_row, style_data_row, style_total_row,
                  font_body, font_caption, font_subheader, ACCENT_POSITIVE, NEUTRAL_600)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = '/home/z/my-project/YMK/Сверка МДК 01.02 М-21.xlsx'

GREEN = Font(name=font_body().name, size=font_body().size, color=ACCENT_POSITIVE, bold=True)
GRAY = Font(name=font_body().name, size=font_body().size, color=NEUTRAL_600, bold=True)

wb = Workbook()
wb.properties.creator = 'Z.ai'


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def status_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = GREEN if '✅' in text else GRAY
    c.alignment = Alignment(horizontal='center', vertical='top')
    return c


def est_height(texts, widths, base=30, per_line=14):
    """Прикидка высоты строки под перенос текста."""
    lines = 1
    for t, w in zip(texts, widths):
        if t:
            lines = max(lines, -(-len(str(t)) // max(int(w / 1.9), 8)))
    return max(base, lines * per_line + 8)


# ============================ Лист 1. Сверка по пунктам ============================
ws = wb.active
ws.title = 'Сверка по пунктам'
setup_sheet(ws, 'Сверка часов: МДК 01.02, группа М-21 (нагрузка → УП → РП → КТП), 2026–2027', last_col=9)
set_widths(ws, {'B': 5, 'C': 27, 'D': 20, 'E': 25, 'F': 24, 'G': 24, 'H': 9, 'I': 44})

h1 = ['№', 'Пункт сверки', 'Учебный план 08.02.09 (2025)', 'Нагрузка 2026–2027 (Барков Д.Н.)',
      'РП ПМ.01 (2024) — справочно', 'КТП М-21 (2026–2027)', 'Статус', 'Разъяснение']
for i, h in enumerate(h1, 2):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 2, 9)

rows1 = [
    ('1', 'Годовой объём уч. часов МДК 01.02', '104 уч. ч', '38 (с.3) + 66 (с.4) = 104 «теор.»',
     '102 уч. ч + 2 ч аттестация = 104', '104 (титул, Т1 «Всего», Т2 строка МДК)', '✅',
     'Все четыре уровня дают 104 ч'),
    ('2', 'В т.ч. в форме практической подготовки (лаб./практ.)', '48', '48 — «Лаб./практ. (в т.ч.)», с.4',
     '48 (кол. 4 строки МДК)', '48 (титул; Σ кол. 4 Т2)', '✅', '48 = 48 = 48 = 48'),
    ('3', 'Теоретическое обучение', '54', 'отдельно не выделено (в составе «теор.»)',
     '54: темы 2.1–2.4 = 12+16+18+8', '54: с.3 = 36, с.4 = 18', '✅',
     '54 ч теории видны в УП, РП и КТП; в ведомости входит в состав «теор.»'),
    ('4', 'Семестр 3', '38', '38 («Теор. часы», столбец МДК 01.02 с.3)', 'в составе годового 102+2',
     '38 = 36 теор + 2 КДЗ (Т1)', '✅', 'Семестровые часы совпадают'),
    ('5', 'Семестр 4', '66', '66 («Теор. часы»), в т.ч. 48 лаб./практ.', 'в составе годового 102+2',
     '66 = 18 теор + 48 практ (Т1)', '✅', 'Семестровые часы совпадают'),
    ('6', 'Аттестация по МДК — комплексный дифференцированный зачёт', '2 (графа «Пром. атт»)',
     'внутри 38 ч «теор.» с.3', '2 (отдельная строка тематического плана)',
     '2 — занятие №52 «Дифференцированный зачет» (Т2)', '✅',
     'Различие только в классификации 2 ч; суммы не меняются: 38 = 38'),
    ('7', 'Консультации (с.4)', 'на уровне модуля ПМ.01: 4 ч (сем. 4)', '2 (строка «Консульт.»)',
     '—', 'в КТП не входит', 'ℹ️',
     'Сверх 104 уч. ч; консультации к экзамену по модулю — см. лист «Разъяснения», п. 6.4'),
    ('8', 'Экзамен по модулю ПМ.01 (с.4)', 'пром. аттестация модуля: 8 ч', '8 (строка «Экзамен»)', '—',
     'в КТП не входит', 'ℹ️',
     'Аттестация по модулю в целом; сверх 104 уч. ч — см. лист «Разъяснения», п. 6.4'),
    ('9', 'Итого по ведомости за год (столбец МДК 01.02)', '104 уч. ч + аттестации', '114 = 38 + (66 + 2 + 8)',
     '104', '104 уч. ч', '✅', '114 = 104 уч. + 2 конс. + 8 экз.; расхождений в уч. часах нет'),
]
r = 5
for idx, row in enumerate(rows1):
    for j, v in enumerate(row, 2):
        ws.cell(row=r, column=j, value=v)
    style_data_row(ws, r, 2, 9, idx)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='top')
    status_cell(ws, r, 8, row[6])
    for col in (3, 9):
        ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    for col in (4, 5, 6, 7):
        ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = est_height(list(row[1:6]) + [row[7]], [27, 20, 25, 24, 24, 44])
    r += 1
ws.freeze_panes = 'C5'

# ============================ Лист 2. Семестры и темы ============================
ws2 = wb.create_sheet('Семестры и темы')
setup_sheet(ws2, 'Распределение по семестрам и построчный перенос тем РП → КТП', last_col=8)
set_widths(ws2, {'B': 34, 'C': 14, 'D': 16, 'E': 16, 'F': 16, 'G': 14, 'H': 9})

ws2.cell(row=4, column=2, value='Семестр')
ws2.cell(row=4, column=3, value='УП (МДК.01.02)')
ws2.cell(row=4, column=4, value='Нагрузка («Теор. часы»)')
ws2.cell(row=4, column=5, value='КТП (Таблица 1)')
ws2.cell(row=4, column=6, value='Состав по КТП')
ws2.cell(row=4, column=7, value='Статус')
style_header_row(ws2, 4, 2, 7)

sem_rows = [
    ('Семестр 3', '38', '38', '38', '36 ч теория (темы 2.1–2.3 нач.) + 2 ч КДЗ', '✅'),
    ('Семестр 4', '66', '66 (в т.ч. 48 лаб./практ.)', '66', '18 ч теория + 48 ч лаб./практ.', '✅'),
]
r = 5
for idx, row in enumerate(sem_rows):
    for j, v in enumerate(row, 2):
        ws2.cell(row=r, column=j, value=v)
    style_data_row(ws2, r, 2, 7, idx)
    status_cell(ws2, r, 7, row[5])
    for col in (2, 6):
        ws2.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws2.row_dimensions[r].height = est_height(list(row), [34, 14, 16, 16, 16])
    r += 1
ws2.cell(row=r, column=2, value='Год')
ws2.cell(row=r, column=3, value='104')
ws2.cell(row=r, column=4, value='104')
ws2.cell(row=r, column=5, value='104')
ws2.cell(row=r, column=6, value='54 теор + 48 практ + 2 КДЗ; 51 занятие + КДЗ №52')
ws2.cell(row=r, column=7, value='✅')
style_total_row(ws2, r, 2, 7)
ws2.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical='top')
status_cell(ws2, r, 7, '✅')
ws2.cell(row=r, column=7).font = GREEN
ws2.row_dimensions[r].height = 34

# --- блок Б: перенос тем ---
r_b = r + 3
ws2.cell(row=r_b, column=2, value='Тема (раздел 2 РП ПМ.01 / Таблица 2 КТП)')
ws2.cell(row=r_b, column=3, value='РП: теория')
ws2.cell(row=r_b, column=4, value='РП: практ. (кол. 4)')
ws2.cell(row=r_b, column=5, value='РП: всего')
ws2.cell(row=r_b, column=6, value='КТП: кол. 3')
ws2.cell(row=r_b, column=7, value='КТП: кол. 4')
ws2.cell(row=r_b, column=8, value='Статус')
style_header_row(ws2, r_b, 2, 8)

topic_rows = [
    ('Тема 2.1. Автоматизированная система контроля и учета электроэнергии', 12, 20, 32, 32, 20,
     '5 лаб. + 5 практ. работ по 2 ч'),
    ('Тема 2.2. Автоматика питающих линий', 16, 12, 28, 28, 12, '6 практ. работ по 2 ч'),
    ('Тема 2.3. Контроль технического состояния многоквартирного дома и инвентаря', 18, 8, 26, 26, 8,
     '4 практ. работы по 2 ч'),
    ('Тема 2.4. Организация проведения расчетов с потребителями и поставщиками ЖКУ', 8, 8, 16, 16, 8,
     '4 практ. работы по 2 ч (№19–22 по РП)'),
    ('Промежуточная аттестация — комплексный дифференцированный зачёт', None, None, 2, 2, None,
     'занятие №52 в КТП'),
]
r = r_b + 1
first_data = r
for idx, row in enumerate(topic_rows):
    name, th, pr, tot, k3, k4, note = row
    vals = [name, th if th is not None else '—', pr if pr is not None else '—', tot, k3,
            k4 if k4 is not None else '—', note]
    for j, v in enumerate(vals, 2):
        ws2.cell(row=r, column=j, value=v)
    ws2.cell(row=r, column=9, value='✅')
    style_data_row(ws2, r, 2, 9, idx)
    status_cell(ws2, r, 9, '✅')
    for col in (2, 8):
        ws2.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    for col in (3, 4, 5, 6, 7):
        ws2.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='top')
    ws2.row_dimensions[r].height = est_height([name, note], [34, 16])
    r += 1
last_data = r - 1
ws2.cell(row=r, column=2, value='Итого по темам (строка МДК РП/КТП: 104 / 48)')
ws2.cell(row=r, column=3, value=f'=SUM(C{first_data}:C{last_data})')
ws2.cell(row=r, column=4, value=f'=SUM(D{first_data}:D{last_data})')
ws2.cell(row=r, column=5, value=f'=SUM(E{first_data}:E{last_data})')
ws2.cell(row=r, column=6, value=f'=SUM(F{first_data}:F{last_data})')
ws2.cell(row=r, column=7, value=f'=SUM(G{first_data}:G{last_data})')
ws2.cell(row=r, column=8, value='102 + 2 КДЗ = 104; кол. 4 = 48')
ws2.cell(row=r, column=9, value='✅')
style_total_row(ws2, r, 2, 9)
status_cell(ws2, r, 9, '✅')
ws2.cell(row=r, column=9).font = GREEN
ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws2.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical='top')
ws2.row_dimensions[r].height = 34
r += 2
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws2.cell(row=r, column=2,
             value='Контрольные суммы Т2 КТП: Σ кол. 3 тем (102) + КДЗ (2) = строка МДК (104); '
                   'Σ кол. 4 тем (48) = кол. 4 строки МДК (48). Построчный перенос из РП — без изменений названий и колонок.')
c.font = font_caption()
c.alignment = Alignment(wrap_text=True, vertical='top')
ws2.row_dimensions[r].height = 30
ws2.freeze_panes = 'C5'

# ============================ Лист 3. Разъяснения ============================
ws3 = wb.create_sheet('Разъяснения')
setup_sheet(ws3, 'Разъяснения по пунктам, требующим внимания', last_col=4)
set_widths(ws3, {'B': 7, 'C': 30, 'D': 96})

for i, h in enumerate(['№', 'Пункт', 'Разъяснение'], 2):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 2, 4)

notes = [
    ('6.1', '«Теор. часы» в ведомости = все учебные занятия',
     'В ведомости по столбцу МДК 01.02 проставлены «Теор. часы»: 38 (с.3) и 66 (с.4) — это все занятия с преподавателем '
     '(теория + лабораторные/практические вместе), а не только теория. Поэтому 38 + 66 = 104 напрямую сопоставляется со '
     'строкой «Уч. зан. 104» учебного плана и строкой МДК 104 в Т2 КТП. Чистая теория (54 ч) в ведомости не выделяется — '
     'она видна в УП, РП и КТП.'),
    ('6.2', 'Два часа комплексного дифзачёта (семестр 3)',
     'В УП 2 ч стоят в графе «Пром. атт» строки МДК.01.02 и входят в семестровое 38; в РП это отдельная строка '
     '«Промежуточная аттестация (комплексный дифференцированный зачет) — 2»; в КТП — отдельное занятие №52 «Дифференцированный '
     'зачет — 2 ч». В ведомости отдельной строки нет: 2 ч находятся внутри 38 ч «теор.» семестра 3. Расхождение только в '
     'классификации, суммы не меняются: 38 = 38.'),
    ('6.3', '«Лаб./практ. (в т.ч.) 48» в семестре 4 — детализация, а не дополнение',
     'Строка «Лаб./практ. (в т.ч.)» в чётных семестрах ведомости раскрывает состав «теор.» часов (66 = 18 теор. + 48 лаб./практ.) '
     'и в столбцовые итоги и ИТОГО не входит — двойной счёт 108 ч устранён в ведомости ранее. Поэтому 48 ч лаб./практ. ведомости = '
     '48 ч «в т.ч. практической подготовки» УП/РП/КТП (титул и Σ кол. 4 Т2).'),
    ('6.4', 'Консультации 2 ч и экзамен 8 ч (с.4) — за рамками 104 ч',
     'В столбец МДК 01.02 ведомости они попали потому, что экзамен по модулю ПМ.01 и консультации к нему проводит преподаватель, '
     'ведущий МДК 01.02. Эти часы: а) не входят в 104 уч. часа МДК; б) в КТП по МДК не включаются (КТП формируется по кол. 3/4 '
     'строки МДК тематического плана РП); в) в УП показаны на уровне модуля: консультации ПМ.01 — 4 ч (по 2 ч на каждый МДК), '
     'пром. аттестация модуля — 16 ч (экзамен 8 + КДЗ двух МДК по 2 + аттестации УП.01 и ПП.01 по 2). Итог ведомости '
     '114 = 104 уч. + 2 конс. + 8 экз. корректен.'),
    ('6.5', 'Титул КТП «в т.ч. в форме практической подготовки — 48 часов»',
     'Заполняется по кол. 4 строки «МДК 01.02 …» тематического плана РП (правило, закреплённое владельцем для всех КТП: титул и '
     'Т2 = кол. 3/4 строки МДК РП). РП даёт 48, КТП переносит 48 — совпадает с «Практ. подг. 48» УП и детализацией ведомости.'),
    ('6.6', 'Нумерация практических работ в РП (номера с пропусками)',
     'В РП сквозная нумерация работ имеет разрыв: тема 2.3 заканчивается работами №12–15, а тема 2.4 начинается с №19 (№16–18 в РП '
     'отсутствуют). На часы это не влияет (в теме 2.4 ровно 4 работы по 2 ч = 8 ч). В КТП занятия пронумерованы сквозно и независимо '
     '(«Практическое занятие №1…№24»), названия работ перенесены из РП полностью.'),
]
r = 5
for idx, (num, name, text) in enumerate(notes):
    ws3.cell(row=r, column=2, value=num)
    ws3.cell(row=r, column=3, value=name)
    ws3.cell(row=r, column=4, value=text)
    style_data_row(ws3, r, 2, 4, idx)
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='top')
    for col in (3, 4):
        ws3.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.row_dimensions[r].height = est_height([name, text], [30, 96], per_line=13)
    r += 1
r += 1
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
c = ws3.cell(row=r, column=2,
             value='ВЫВОД: часы МДК 01.02 (М-21) полностью синхронизированы: УП 104/48 = нагрузка 104 уч. ч (38+66, плюс 2 конс. и '
                   '8 экз. сверх уч. ч) = РП 102+2/48 = КТП 104/48 (построчный перенос тем 2.1–2.4, контрольные суммы сходятся). '
                   'КТП KTP/КТП МДК 01.02 М-21.docx корректно и может использоваться в работе.')
c.font = font_subheader()
c.alignment = Alignment(wrap_text=True, vertical='top')
ws3.row_dimensions[r].height = 60

wb.save(OUT)
print('OK:', OUT)
